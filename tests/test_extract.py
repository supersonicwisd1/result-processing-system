import pypdf
from docx import Document
from openpyxl import load_workbook
import csv

def extract_xls_data(filepath):
    workbook = load_workbook(filepath)
    sheet = workbook.active

    for row in sheet.iter_rows(values_only=True):
        print(row)

# extract_xls_data("/Users/kene/Desktop/Customer Master Data-Base.xlsx")

def extract_docx_data(filepath):
    doc = Document(filepath)
    full_text = []
    for para in doc.paragraphs:
        full_text.append(para.text)
    print("\n".join(full_text))

# extract_docx_data("/Users/kene/Desktop/Resumes&profile_pics/old/Kenechukwu Orjiene Cover letter.docx")

def extract_pdf_data(filepath):
    with open(filepath, 'rb') as file:
        reader = pypdf.PdfReader(file)
        text = ""
        for page in reader.pages:
            text += page.extract_text() + "\n"
    print(text)

# extract_pdf_data("/Users/kene/Desktop/tutorial/Python Programming .pdf")

def extract_csv_data(filepath):
    extracted_data = []
    with open(filepath, newline='') as csvfile:
        reader = csv.DictReader(csvfile)
        for row in reader:
            # extracted_data.append({
            #     "registration_number": row.get('registration_number'),
            #     "student_name": row.get('student_name'),
            #     "ca_score": float(row.get('ca_score', 0)),
            #     "exam_score": float(row.get('exam_score', 0)),
            #     "total_score": float(row.get('total_score', 0)),
            #     "grade": row.get('grade')
            # })
            print(extracted_data)
extract_csv_data()